import os
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from flask import Flask, request, send_file, render_template, jsonify
from werkzeug.utils import secure_filename
import logging
from datetime import datetime

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuração do Flask
app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}

# Criar diretórios necessários
for folder in [UPLOAD_FOLDER, 'logs']:
    if not os.path.exists(folder):
        os.makedirs(folder)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limite de 16MB

def allowed_file(filename):
    """Verifica se a extensão do arquivo é permitida."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_transactions_from_pdf(pdf_path):
    """
    Extrai as transações do PDF no formato especificado.
    Retorna uma lista de dicionários com as transações.
    """
    transactions = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                lines = text.split('\n')

                # Padrão ajustado para o formato específico
                transaction_pattern = re.compile(
                    r'(\d{2}-\d{2}-\d{4})\s+'    # Data
                    r'(.*?)\s+'                   # Descrição
                    r'(\d{11})\s+'                # ID da operação
                    r'R\$\s*(-?\d+[.,]\d{2})\s+'  # Valor
                    r'R\$\s*\d+[.,]\d{2}'         # Saldo (não capturado)
                )

                current_transaction = None
                for line in lines:
                    line = line.strip()

                    # Pular linhas de cabeçalho e rodapé
                    if any(header in line for header in ['EXTRATO DE CONTA', 'DETALHE DOS MOVIMENTOS', 'Data de geração:', 'Saldo inicial:', 'Saldo final:']):
                        continue

                    # Tentar encontrar uma transação completa
                    match = transaction_pattern.search(line)
                    if match:
                        date, description, operation_id, value = match.groups()

                        # Limpar e converter o valor
                        value = value.replace('.', '').replace(',', '.')
                        value = float(value)

                        transactions.append({
                            'Data': date,
                            'Descrição': description.strip(),
                            'Valor': value
                        })
                        current_transaction = None
                    elif current_transaction is not None:
                        # Adicionar linha à descrição da transação atual
                        transactions[-1]['Descrição'] += ' ' + line.strip()

        logger.info(f"Extraídas {len(transactions)} transações do PDF")
        return transactions
    except Exception as e:
        logger.error(f"Erro ao processar o PDF {pdf_path}: {str(e)}")
        raise

def validate_transactions(transactions):
    """
    Valida se todas as transações foram extraídas corretamente.
    """
    if not transactions:
        logger.error("Nenhuma transação encontrada")
        return False

    for idx, transaction in enumerate(transactions):
        if not all(key in transaction for key in ['Data', 'Descrição', 'Valor']):
            logger.error(f"Transação {idx} está faltando campos obrigatórios")
            return False

        try:
            # Validar formato da data
            datetime.strptime(transaction['Data'], '%d-%m-%Y')

            # Validar valor
            if not isinstance(transaction['Valor'], (int, float)):
                raise ValueError(f"Valor inválido na transação {idx}")

            # Validar descrição
            if not transaction['Descrição'].strip():
                raise ValueError(f"Descrição vazia na transação {idx}")

        except Exception as e:
            logger.error(f"Erro na validação da transação {idx}: {str(e)}")
            return False

    return True

def format_excel(excel_path):
    """
    Aplica formatação avançada ao arquivo Excel.
    Inclui cores, fontes e alinhamento.
    """
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Estilos para o cabeçalho
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Formatar cabeçalho
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Formatar células de dados
        for row in sheet.iter_rows(min_row=2):
            # Data
            row[0].alignment = Alignment(horizontal='center')
            # Descrição
            row[1].alignment = Alignment(horizontal='left')
            # Valor
            row[2].alignment = Alignment(horizontal='right')
            if row[2].value:
                valor = float(str(row[2].value).replace('.', '').replace(',', '.'))
                if valor < 0:
                    row[2].font = Font(color="FF0000")  # Vermelho para valores negativos
                else:
                    row[2].font = Font(color="008000")  # Verde para valores positivos

        # Ajustar largura das colunas
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_path)
        workbook.close()
        logger.info(f"Excel formatado com sucesso: {excel_path}")
    except Exception as e:
        logger.error(f"Erro ao formatar o Excel: {str(e)}")

def create_excel(transactions, excel_path):
    """
    Cria arquivo Excel com as transações e aplica formatação.
    """
    try:
        df = pd.DataFrame(transactions)
        # Ordenar por data
        df['Data'] = pd.to_datetime(df['Data'], format='%d-%m-%Y')
        df = df.sort_values('Data')
        df['Data'] = df['Data'].dt.strftime('%d-%m-%Y')

        # Formatar valores
        df['Valor'] = df['Valor'].map(lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))

        # Salvar Excel
        df.to_excel(excel_path, sheet_name='Extrato', index=False)

        # Aplicar formatação
        format_excel(excel_path)
        logger.info(f"Arquivo Excel criado com sucesso: {excel_path}")
        return True
    except Exception as e:
        logger.error(f"Erro ao criar arquivo Excel: {str(e)}")
        return False

@app.route('/', methods=['GET'])
def index():
    """Rota principal que renderiza a página inicial."""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Processa o upload do arquivo PDF e retorna o Excel convertido.
    """
    try:
        if 'file' not in request.files:
            return 'Nenhum arquivo enviado', 400

        file = request.files['file']
        if file.filename == '':
            return 'Nenhum arquivo selecionado', 400

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{timestamp}_{filename}")
            file.save(pdf_path)

            # Extrair transações
            try:
                transactions = extract_transactions_from_pdf(pdf_path)

                # Validar transações
                if not validate_transactions(transactions):
                    return 'Erro na validação das transações', 400

                # Criar Excel
                excel_filename = f"{timestamp}_{filename.replace('.pdf', '.xlsx')}"
                excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
                if not create_excel(transactions, excel_path):
                    return 'Erro ao criar arquivo Excel', 400

                # Limpar arquivo PDF
                try:
                    os.remove(pdf_path)
                except Exception as e:
                    logger.error(f"Erro ao excluir PDF: {str(e)}")

                # Enviar Excel
                return send_file(
                    excel_path,
                    as_attachment=True,
                    download_name=filename.replace('.pdf', '.xlsx'),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                logger.error(f"Erro ao processar arquivo: {str(e)}")
                return f'Erro ao processar arquivo: {str(e)}', 400

    except Exception as e:
        logger.error(f"Erro no upload: {str(e)}")
        return 'Erro interno do servidor', 500

@app.errorhandler(413)
def too_large(e):
    """Manipula erro de arquivo muito grande."""
    return 'Arquivo muito grande. Tamanho máximo permitido é 16MB', 413

@app.errorhandler(500)
def internal_error(e):
    """Manipula erro interno do servidor."""
    return 'Erro interno do servidor', 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
